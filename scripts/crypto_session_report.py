"""
Build a 2-sheet Excel report (Executive Summary + Trade Details) for a crypto
paper-trading session, mirroring the F&O session report format.

Currency is USDT/USD — every *_USDT pair is quoted against dollar-pegged Tether,
so there is NO rupee figure anywhere in a crypto run.

    PYTHONPATH=. python scripts/crypto_session_report.py \
        --session-dir storage/crypto_live [--out <path.xlsx>]
"""
import argparse
import json
import pathlib
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEAD = PatternFill("solid", fgColor="1F3864")
SUBHEAD = PatternFill("solid", fgColor="2E5496")
KPI = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _load(session_dir):
    d = pathlib.Path(session_dir)
    trades = [json.loads(x) for x in (d / "paper_trades.jsonl").read_text(encoding="utf-8").splitlines() if x.strip()]
    return trades


def _round_trips(trades):
    """Pair each exit with the most recent open entry of the same symbol (FIFO)."""
    open_entries = {}
    rows = []
    for t in trades:
        sym = t.get("symbol")
        if t.get("type") == "entry":
            open_entries.setdefault(sym, []).append(t)
        elif t.get("type") == "exit":
            entry = open_entries.get(sym, []).pop(0) if open_entries.get(sym) else None
            if not entry:
                continue
            qty = entry.get("quantity") or 1
            buy = entry["price"]
            sell = t["price"]
            capital = buy * qty
            revenue = sell * qty
            net = revenue - capital
            t0 = datetime.fromisoformat(entry["timestamp"])
            t1 = datetime.fromisoformat(t["timestamp"])
            hold_s = (t1 - t0).total_seconds()
            gate = entry.get("gate_evaluation") or {}
            rows.append({
                "entry_ts": t0.strftime("%H:%M:%S"), "exit_ts": t1.strftime("%H:%M:%S"),
                "symbol": sym, "strategy": "crypto_lag", "direction": "LONG",
                "reference": entry.get("lag_reference"), "qty": qty,
                "buy": buy, "sell": sell, "capital": capital, "revenue": revenue,
                "net": net, "net_pct": net / capital if capital else 0,
                "hold_min": hold_s / 60.0,
                "expected_net_pct": gate.get("net_profit_pct"),
                "expected_annualized_pct": gate.get("annualized_return_pct"),
            })
    return rows


def build(session_dir, out=None):
    trades = _load(session_dir)
    rows = _round_trips(trades)
    summary = next((t for t in reversed(trades) if t.get("type") == "session_summary"), None)
    stats = (summary or {}).get("stats", {})
    open_positions = (summary or {}).get("open_positions", {})
    date_str = datetime.now().strftime("%B %d, %Y")

    wb = openpyxl.Workbook()

    # ---------------- Trade Details ----------------
    td = wb.active
    td.title = "Trade Details"
    cols = ["Entry Time", "Exit Time", "Symbol", "Strategy", "Direction", "Lag Reference",
            "Quantity", "Buy Price (USDT)", "Sell Price (USDT)", "Capital Deployed (USDT)",
            "Gross Revenue (USDT)", "Net PnL (USDT)", "Net PnL %", "Hold (min)",
            "Entry Expected Net %", "Entry Expected Annualized %"]
    td.append(cols)
    for c in td[1]:
        c.fill = SUBHEAD; c.font = WHITE; c.border = THIN; c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        td.append([r["entry_ts"], r["exit_ts"], r["symbol"], r["strategy"], r["direction"],
                   r["reference"], r["qty"], r["buy"], r["sell"], r["capital"], r["revenue"],
                   r["net"], r["net_pct"], r["hold_min"], r["expected_net_pct"], r["expected_annualized_pct"]])
    last = td.max_row
    if rows:
        td.append(["Total / Avg", None, None, None, None, None,
                   f"=SUM(G2:G{last})", None, None, f"=SUM(J2:J{last})", f"=SUM(K2:K{last})",
                   f"=SUM(L2:L{last})", f"=AVERAGE(M2:M{last})", f"=AVERAGE(N2:N{last})", None, None])
        for c in td[td.max_row]:
            c.font = BOLD; c.fill = KPI
    # number formats
    for row in td.iter_rows(min_row=2, max_row=td.max_row):
        for c in row:
            if c.column_letter in ("H", "I", "J", "K", "L"):
                c.number_format = "#,##0.00"
            elif c.column_letter in ("M", "O", "P"):
                c.number_format = "0.0000%" if c.column_letter == "M" else "0.000"
            elif c.column_letter == "N":
                c.number_format = "0.0"
    widths = [10, 10, 12, 12, 10, 13, 9, 15, 15, 18, 18, 14, 12, 10, 16, 20]
    for i, w in enumerate(widths, 1):
        td.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---------------- Executive Summary ----------------
    es = wb.create_sheet("Executive Summary", 0)
    es["A1"] = "MARKET TERMINAL - CRYPTO SESSION REPORT"
    es["A1"].font = Font(size=14, bold=True, color="1F3864")
    es["A3"] = f"Session Date: {date_str}  |  Feed: crypto.com (live)  |  Mode: Live Paper Trading  |  Currency: USDT (USD-pegged)"
    es["A3"].font = Font(italic=True, color="595959")

    total_net = sum(r["net"] for r in rows)
    total_capital = sum(r["capital"] for r in rows)
    total_turnover = sum(r["capital"] + r["revenue"] for r in rows)
    kpis = [
        ("TOTAL CAPITAL DEPLOYED (USDT)", total_capital),
        ("TOTAL TURNOVER (USDT)", total_turnover),
        ("NET PnL (USDT)", total_net),
        ("RETURN ON DEPLOYED CAPITAL", (total_net / total_capital) if total_capital else 0),
    ]
    for i, (label, val) in enumerate(kpis):
        col = 1 + i * 2
        lc = es.cell(row=5, column=col, value=label); lc.fill = HEAD; lc.font = WHITE; lc.alignment = Alignment(wrap_text=True, horizontal="center")
        vc = es.cell(row=6, column=col, value=val); vc.fill = KPI; vc.font = Font(size=12, bold=True); vc.alignment = Alignment(horizontal="center")
        vc.number_format = "0.00%" if i == 3 else "#,##0.00"

    es["A8"] = "Session Search & Validation Summary"; es["A8"].font = BOLD; es["A8"].fill = SUBHEAD; es["A8"].font = WHITE
    es.append([])  # spacer already at 8
    val_rows = [
        ("Metric", "Value", "Description"),
        ("Ticks Processed", stats.get("ticks"), "Price ticks polled from the live crypto.com feed."),
        ("Quotes Received", stats.get("quotes_received"), "Individual instrument quotes ingested."),
        ("Lag Signals Detected", stats.get("lag_signals"), "Reference/target lag conditions observed."),
        ("Entries Taken", stats.get("entries"), "Signals that cleared the cost + return gate."),
        ("Exits Taken", stats.get("exits"), "Positions closed (take-profit / stop-loss / duration)."),
        ("Blocked by Ranking Engine", stats.get("blocked"), "Signals rejected: not profitable after costs / below return hurdle."),
        ("Poll Errors", stats.get("poll_errors"), "Failed feed polls during the session."),
        ("Open Positions at Close", ", ".join(open_positions) or "none", "Positions still held when the session ended."),
    ]
    for j, vr in enumerate(val_rows):
        es.append(vr)
        if j == 0:
            for c in es[es.max_row]:
                c.font = WHITE; c.fill = SUBHEAD
    es.column_dimensions["A"].width = 28
    es.column_dimensions["C"].width = 30
    for col in ("B", "D", "E", "F", "G", "H"):
        es.column_dimensions[col].width = 16

    out = out or str(pathlib.Path(session_dir) / f"crypto_session_report_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(out)
    return out, total_net, len(rows)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--session-dir", default="storage/crypto_live")
    p.add_argument("--out", default=None)
    args = p.parse_args()
    path, net, n = build(args.session_dir, args.out)
    print(f"Wrote {path}  ({n} round-trip trades, net {net:+.2f} USDT)")


if __name__ == "__main__":
    main()
