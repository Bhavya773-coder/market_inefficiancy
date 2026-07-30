"""
Build a 2-sheet Excel report (Executive Summary + Trade Details) for an F&O /
spot inefficiency session, mirroring the 2026-07-21 report layout.

Currency is INR (NSE / BSE / MCX). Per-leg prices, quantities and round-trip
costs are NOT logged by the current engine, so those columns are omitted rather
than fabricated; Capital Deployed is derived exactly as net_profit / net_profit_pct.

    PYTHONPATH=. python scripts/fno_session_report.py \
        --session-dir storage/session_live_20260723 [--out <path.xlsx>]
"""
import argparse
import json
import pathlib
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEAD = PatternFill("solid", fgColor="1F3864")
SUBHEAD = PatternFill("solid", fgColor="2E5496")
KPI = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
STRATEGIES = ["nse_bse_arb", "futures_basis", "put_call_parity", "mcx_calendar"]


# How each leg pair is actually funded in a real account. Leverage is only
# honest for the first category; the other two need cash or borrowed stock,
# which a margin multiplier does not provide.
MARGINED = "MARGINED"            # futures vs futures - real F&O margin applies
CASH_REQUIRED = "CASH REQUIRED"  # a leg buys actual shares - no leverage
SHORT_SPOT = "SHORT SPOT"        # a leg shorts actual shares - needs SLB

FUNDING = {
    ("mcx_calendar", "CASH_AND_CARRY"): MARGINED,
    ("mcx_calendar", "REVERSE_CASH_AND_CARRY"): MARGINED,
    ("futures_basis", "CASH_AND_CARRY"): CASH_REQUIRED,
    ("futures_basis", "REVERSE_CASH_AND_CARRY"): SHORT_SPOT,
    ("put_call_parity", "CONVERSION"): CASH_REQUIRED,
    ("put_call_parity", "REVERSAL"): SHORT_SPOT,
}
FUNDING_NOTE = {
    MARGINED: "Both legs are futures. Exchange margin applies and is a "
              "fraction of notional, so leverage here is real - arguably "
              "conservative, since offsetting calendar legs get margin relief.",
    CASH_REQUIRED: "One leg buys the actual shares. 1L of cash cannot buy 6L "
                   "of stock; F&O margin does not fund an equity purchase. "
                   "Leverage on this row is fictional.",
    SHORT_SPOT: "One leg shorts the actual shares. Indian retail needs SLB "
                "for an overnight short, which is thin and expensive. Both "
                "the leverage AND the executability are questionable.",
}


def funding_basis(strategy, direction):
    if strategy == "nse_bse_arb":
        return CASH_REQUIRED  # buy on one exchange, sell on the other
    return FUNDING.get((strategy, direction), CASH_REQUIRED)


def _load(session_dir):
    d = pathlib.Path(session_dir)
    rows = [json.loads(x) for x in (d / "inefficiencies.jsonl").read_text(encoding="utf-8").splitlines() if x.strip()]
    trades = [json.loads(x) for x in (d / "paper_trades.jsonl").read_text(encoding="utf-8").splitlines() if x.strip()]
    return rows, trades


def _enrich_captures(rows, trades):
    by_id = {}
    for r in rows:
        by_id.setdefault(r["opportunity_id"], []).append(r)
    out = []
    for c in trades:
        if c.get("type") != "capture":
            continue
        cand = by_id.get(c["opportunity_id"], [])
        m = min(cand, key=lambda r: abs(r["net_profit"] - c["net_profit"])) if cand else {}
        net = c["net_profit"]
        net_pct = m.get("net_profit_pct")
        capital = net / (net_pct / 100) if net_pct else None
        direction = m.get("direction")
        out.append({
            "time": c["timestamp"][11:19], "asset": c["asset"], "strategy": c["strategy"],
            "direction": direction, "action": m.get("action"),
            "net": net, "net_pct_frac": (net_pct / 100) if net_pct is not None else None,
            "ann_frac": (m.get("annualized_return_pct") / 100) if m.get("annualized_return_pct") is not None else None,
            "capital": capital,
            "funding": funding_basis(c["strategy"], direction),
        })
    return out


def _session_config(trades):
    """Funding basis the run was started with, if the runner recorded it."""
    cfg = next((t for t in trades if t.get("type") == "session_config"), None)
    if not cfg:
        return None
    return {"capital": cfg.get("capital"), "leverage": cfg.get("leverage", 1.0),
            "buying_power": cfg.get("buying_power")}


def build(session_dir, out=None):
    rows, trades = _load(session_dir)
    cfg = _session_config(trades)
    caps = _enrich_captures(rows, trades)
    date_str = datetime.now().strftime("%B %d, %Y")

    wb = openpyxl.Workbook()

    # ---------------- Trade Details ----------------
    td = wb.active
    td.title = "Trade Details"
    cols = ["Time", "Asset", "Strategy", "Direction", "Action (Buy/Sell)",
            "Net Profit (INR)", "Net Profit %", "Annualized Return %",
            "Capital Deployed (INR)", "Funding basis", "Real at cash capital?"]
    td.append(cols)
    for c in td[1]:
        c.fill = SUBHEAD; c.font = WHITE; c.alignment = Alignment(wrap_text=True, vertical="center")
    cash_cap = (cfg or {}).get("capital")
    for r in caps:
        if r["funding"] == MARGINED:
            real = "YES (margined)"
        elif cash_cap and r["capital"] and r["capital"] <= cash_cap:
            real = "YES (fits cash)"
        else:
            real = "NO"
        td.append([r["time"], r["asset"], r["strategy"], r["direction"], r["action"],
                   r["net"], r["net_pct_frac"], r["ann_frac"], r["capital"],
                   r["funding"], real])
    last = td.max_row
    n_caps = len(caps)
    total_net_col = sum(r["net"] for r in caps)
    avg_net_pct = sum(r["net_pct_frac"] for r in caps if r["net_pct_frac"] is not None) / n_caps if n_caps else 0
    avg_ann_pct = sum(r["ann_frac"] for r in caps if r["ann_frac"] is not None) / n_caps if n_caps else 0
    total_capital_col = sum(r["capital"] for r in caps if r["capital"])
    td.append(["Total / Avg", None, None, None, None,
               total_net_col, avg_net_pct, avg_ann_pct, total_capital_col])
    for c in td[td.max_row]:
        c.font = BOLD; c.fill = KPI
    for row in td.iter_rows(min_row=2, max_row=td.max_row):
        for c in row:
            if c.column_letter in ("F", "I"):
                c.number_format = "#,##0.00"
            elif c.column_letter in ("G", "H"):
                c.number_format = "0.0000%"
    for i, w in enumerate([10, 12, 16, 24, 30, 15, 12, 16, 20, 16, 20], 1):
        td.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---------------- Executive Summary ----------------
    es = wb.create_sheet("Executive Summary", 0)
    es["A1"] = "MARKET TERMINAL - F&O / SPOT INEFFICIENCY SESSION REPORT"
    es["A1"].font = Font(size=14, bold=True, color="1F3864")
    es["A3"] = (f"Session Date: {date_str}  |  Feed: Dhan (live NSE / BSE / MCX)  |  "
                f"Mode: Live Paper Trading  |  Currency: INR")
    es["A3"].font = Font(italic=True, color="595959")

    total_net = total_net_col
    total_capital = total_capital_col
    kpis = [
        ("TOTAL CAPITAL DEPLOYED (INR)", "#,##0.00", total_capital),
        ("TOTAL NET PROFIT (INR)", "#,##0.00", total_net),
        ("RETURN ON DEPLOYED CAPITAL", "0.000%", (total_net / total_capital) if total_capital else 0),
        ("EXECUTABLE CAPTURES", "0", len(caps)),
    ]
    for i, (label, fmt, val) in enumerate(kpis):
        col = 1 + i * 2
        lc = es.cell(row=5, column=col, value=label); lc.fill = HEAD; lc.font = WHITE
        lc.alignment = Alignment(wrap_text=True, horizontal="center")
        vc = es.cell(row=6, column=col, value=val); vc.fill = KPI; vc.font = Font(size=12, bold=True)
        vc.alignment = Alignment(horizontal="center"); vc.number_format = fmt

    # ---- Funding-reality split -------------------------------------
    # A leverage multiplier raises buying power uniformly, but only
    # futures-vs-futures legs are genuinely margined. Splitting the P&L keeps
    # a levered run from reading as if all of it were executable.
    if cfg and cfg.get("leverage", 1.0) > 1:
        es.append(["Funding Reality Split (leverage applied)"])
        es[es.max_row][0].font = WHITE
        es[es.max_row][0].fill = SUBHEAD
        es.append([f"Cash capital {cfg['capital']:,.0f} INR  x  leverage "
                   f"{cfg['leverage']:g}  =  buying power "
                   f"{cfg['buying_power']:,.0f} INR"])
        es[es.max_row][0].font = Font(italic=True, color="595959")
        es.append(["Funding basis", "Captures", "Net P&L (INR)",
                   "% of total", "What it means"])
        for c in es[es.max_row]:
            c.font = WHITE; c.fill = SUBHEAD
        order = [MARGINED, CASH_REQUIRED, SHORT_SPOT]
        by_fund = {k: {"n": 0, "net": 0.0} for k in order}
        for r in caps:
            by_fund[r["funding"]]["n"] += 1
            by_fund[r["funding"]]["net"] += r["net"]
        for k in order:
            v = by_fund[k]
            es.append([k, v["n"], v["net"],
                       (v["net"] / total_net) if total_net else 0,
                       FUNDING_NOTE[k]])
            es.cell(row=es.max_row, column=3).number_format = "#,##0.00"
            es.cell(row=es.max_row, column=4).number_format = "0.0%"
            es.cell(row=es.max_row, column=5).alignment = Alignment(
                wrap_text=True, vertical="top")
        real = by_fund[MARGINED]["net"]
        fake = by_fund[CASH_REQUIRED]["net"] + by_fund[SHORT_SPOT]["net"]
        es.append([])
        es.append(["EXECUTABLE AT THIS CASH LEVEL", real])
        es[es.max_row][0].font = Font(bold=True, color="006100")
        es.cell(row=es.max_row, column=2).number_format = "#,##0.00"
        es.append(["REQUIRES CASH / STOCK BORROW YOU DO NOT HAVE", fake])
        es[es.max_row][0].font = Font(bold=True, color="C00000")
        es.cell(row=es.max_row, column=2).number_format = "#,##0.00"
        es.append([])

    # Performance breakdown by strategy (computed directly in Python — no
    # formula cells, so the numbers show up even in viewers that don't
    # recalculate on open)
    es.append(["Performance Breakdown by Strategy"])
    es[es.max_row][0].font = WHITE
    es[es.max_row][0].fill = SUBHEAD
    es.append(["Strategy", "Trades Taken", "Capital Deployed", "Net Profit", "Return on Capital"])
    for c in es[es.max_row]:
        c.font = WHITE; c.fill = SUBHEAD
    # Derived, not hardcoded: the funding-split block above is optional, so a
    # fixed row number silently mis-sums the totals when it is present.
    first = es.max_row + 1
    present_strategies = STRATEGIES + sorted({r["strategy"] for r in caps} - set(STRATEGIES))
    strat_totals = {}
    for strat in present_strategies:
        rows_for = [r for r in caps if r["strategy"] == strat]
        n_trades = len(rows_for)
        cap_dep = sum(r["capital"] for r in rows_for if r["capital"])
        net_prof = sum(r["net"] for r in rows_for)
        strat_totals[strat] = (n_trades, cap_dep, net_prof)
        es.append([strat, n_trades, cap_dep, net_prof, (net_prof / cap_dep) if cap_dep else 0])
    tot_trades = sum(v[0] for v in strat_totals.values())
    tot_cap = sum(v[1] for v in strat_totals.values())
    tot_net = sum(v[2] for v in strat_totals.values())
    es.append(["Total", tot_trades, tot_cap, tot_net, (tot_net / tot_cap) if tot_cap else 0])
    for c in es[es.max_row]:
        c.font = BOLD
    for r in range(first, es.max_row + 1):
        es.cell(row=r, column=3).number_format = "#,##0.00"
        es.cell(row=r, column=4).number_format = "#,##0.00"
        es.cell(row=r, column=5).number_format = "0.000%"

    # Validation summary
    total = len(rows)
    execu = sum(1 for r in rows if r.get("is_executable"))
    rc = Counter()
    for r in rows:
        for reason in (r.get("rejection_reasons") or []):
            rc[reason] += 1
    es.append([])
    es.append(["Session Search & Validation Summary"]); es[es.max_row][0].font = WHITE; es[es.max_row][0].fill = SUBHEAD
    es.append(["Metric", "Value", "Description"])
    for c in es[es.max_row]:
        c.font = WHITE; c.fill = SUBHEAD
    val_rows = [
        ("Total Opportunities Scanned", total, "Leg-pairs evaluated from the live Dhan feed this session."),
        ("Total Executable Signals", execu, "Passed all cost, liquidity and capital filters."),
        ("Total Rejected Signals", total - execu, "Failed at least one hurdle."),
        ("  - Below Return Hurdle", rc.get("below_min_annualized_return", 0), "Failed the minimum annualized ROC threshold."),
        ("  - Non-Profitable after Costs", rc.get("not_profitable_after_round_trip_costs", 0), "Gross spread insufficient to cover brokerage, taxes, slippage."),
        ("  - Blocked by Insufficient Capital", rc.get("insufficient_capital", 0), "Required capital exceeded the available limit."),
    ]
    for vr in val_rows:
        es.append(vr)
    es.append([])
    es.append(["Note: per-leg buy/sell prices, quantities and round-trip costs are not logged by the "
               "current engine, so those columns are omitted. Capital Deployed = Net Profit / Net Profit %."])
    es[es.max_row][0].font = Font(italic=True, color="808080")

    es.column_dimensions["A"].width = 32
    es.column_dimensions["C"].width = 30
    for col in ("B", "D", "E", "F", "G", "H"):
        es.column_dimensions[col].width = 16

    out = out or str(pathlib.Path(session_dir) / f"fno_session_report_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(out)
    return out, total_net, len(caps)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--session-dir", default="storage/session_live_20260723")
    p.add_argument("--out", default=None)
    args = p.parse_args()
    path, net, n = build(args.session_dir, args.out)
    print(f"Wrote {path}  ({n} captures, net {net:+.2f} INR)")


if __name__ == "__main__":
    main()
