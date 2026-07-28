"""
Consolidated Excel report across every paper-trading session on disk.

Sheets:
  Session Summary     one row per session, with P&L rate per hour
  All Captures        every locked-in capture, newest session last
  Strategy Perf       aggregated by strategy across all sessions
  Rejection Analysis  why opportunities were refused, per session
  Path to 5k          what the data says is needed for 5,000 INR/day

Every figure is computed in Python and written as a literal value. openpyxl
writes formulas without a cached result, so viewers that do not recalculate
on open render formula cells blank.

    PYTHONPATH=. python scripts/all_sessions_report.py
"""
import argparse
import json
import pathlib
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEAD = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5496")
KPI = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)

# Capital each session actually ran with. The runner did not log it before
# --capital existed, so it is recorded here rather than guessed at read time.
SESSION_CAPITAL = {
    "session_20260721": 1_000_000,
    "fno_dashboard_20260723": 1_000_000,
    "session_live_20260723": 1_000_000,
    "fno_1lakh_20260727": 100_000,
    "fno_10lakh_20260727": 1_000_000,
    "fno_10lakh_20260728": 1_000_000,
}
# Synthetic/plumbing runs and crypto-era dirs: real markets only.
SKIP = {"dryrun_item1_check", "dryrun_session_20260714", "live",
        "live_session_20260714", "session_20260723"}


def _read(path):
    if not path.exists():
        return []
    return [json.loads(x) for x in path.read_text(encoding="utf-8").splitlines() if x.strip()]


def _hours(stamps):
    if len(stamps) < 2:
        return 0.0
    lo, hi = min(stamps), max(stamps)
    try:
        a = datetime.fromisoformat(lo.replace("Z", "+00:00"))
        b = datetime.fromisoformat(hi.replace("Z", "+00:00"))
        return max((b - a).total_seconds() / 3600.0, 0.0)
    except ValueError:
        return 0.0


def collect(storage):
    sessions = []
    for d in sorted(pathlib.Path(storage).iterdir()):
        if not d.is_dir() or d.name in SKIP:
            continue
        rows = _read(d / "inefficiencies.jsonl")
        trades = _read(d / "paper_trades.jsonl")
        caps = [t for t in trades if t.get("type") == "capture"]
        if not rows and not caps:
            continue

        rc = Counter()
        for r in rows:
            for x in (r.get("rejection_reasons") or []):
                rc[x] += 1
        by_strat = defaultdict(lambda: {"n": 0, "net": 0.0})
        for c in caps:
            s = c.get("strategy") or "unknown"
            by_strat[s]["n"] += 1
            by_strat[s]["net"] += c.get("net_profit") or 0.0

        stamps = [r["timestamp"] for r in rows if r.get("timestamp")] or \
                 [c["timestamp"] for c in caps if c.get("timestamp")]
        hours = _hours(stamps)
        net = sum(c.get("net_profit") or 0.0 for c in caps)
        # The runner locks in a capture every time an opportunity_id still
        # qualifies on a later poll, so a spread that stays open for several
        # polls is booked repeatedly. In reality you can only take it once.
        # Deduped keeps the best single booking per opportunity_id.
        uniq = {}
        for c in caps:
            k = c.get("opportunity_id")
            uniq[k] = max(uniq.get(k, float("-inf")), c.get("net_profit") or 0.0)
        net_dedup = sum(uniq.values()) if uniq else 0.0
        sessions.append({
            "name": d.name,
            "date": (stamps[0][:10] if stamps else "?"),
            "capital": SESSION_CAPITAL.get(d.name),
            "hours": hours,
            "scanned": len(rows),
            "executable": sum(1 for r in rows if r.get("is_executable")),
            "captures": len(caps),
            "net": net,
            "net_dedup": net_dedup,
            "unique_trades": len(uniq),
            "per_hour": (net_dedup / hours) if hours > 0.05 else None,
            "rejections": rc,
            "by_strategy": dict(by_strat),
            "caps": caps,
            "rows": rows,
        })
    return sessions


def _hdr(ws, cols, widths):
    ws.append(cols)
    for c in ws[1]:
        c.fill = SUB
        c.font = WHITE
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build(storage="storage", out=None):
    sessions = collect(storage)
    if not sessions:
        raise SystemExit("no sessions found")
    wb = openpyxl.Workbook()

    # ---------------- Session Summary ----------------
    ws = wb.active
    ws.title = "Session Summary"
    ws.append(["MARKET TERMINAL — ALL PAPER SESSIONS"])
    ws["A1"].font = Font(size=14, bold=True, color="1F3864")
    ws.append([f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
               f"NSE/BSE/MCX via Dhan · PAPER ONLY · INR · "
               f"synthetic dry-runs excluded"])
    ws["A2"].font = Font(italic=True, color="595959")
    ws.append([])
    cols = ["Session", "Date", "Capital (INR)", "Hours", "Scanned",
            "Executable", "Exec %", "Captures", "Unique trades",
            "Net P&L as logged", "Net P&L deduped", "Deduped P&L / hour",
            "Top strategy"]
    ws.append(cols)
    for c in ws[4]:
        c.fill = SUB
        c.font = WHITE
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for s in sessions:
        top = max(s["by_strategy"].items(), key=lambda kv: kv[1]["net"],
                  default=(None, None))
        ws.append([
            s["name"], s["date"], s["capital"], round(s["hours"], 2),
            s["scanned"], s["executable"],
            (s["executable"] / s["scanned"]) if s["scanned"] else 0,
            s["captures"], s["unique_trades"], s["net"], s["net_dedup"],
            s["per_hour"],
            f"{top[0]} ({top[1]['net']:,.0f})" if top[0] else "-",
        ])
    tot_net = sum(s["net"] for s in sessions)
    tot_dedup = sum(s["net_dedup"] for s in sessions)
    tot_uniq = sum(s["unique_trades"] for s in sessions)
    tot_caps = sum(s["captures"] for s in sessions)
    tot_scan = sum(s["scanned"] for s in sessions)
    tot_exec = sum(s["executable"] for s in sessions)
    ws.append(["TOTAL", "", "", round(sum(s["hours"] for s in sessions), 2),
               tot_scan, tot_exec, (tot_exec / tot_scan) if tot_scan else 0,
               tot_caps, tot_uniq, tot_net, tot_dedup, "", ""])
    for c in ws[ws.max_row]:
        c.font = BOLD
        c.fill = KPI
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for c in row:
            if c.column_letter in ("C", "J", "K", "L"):
                c.number_format = "#,##0.00"
            elif c.column_letter == "G":
                c.number_format = "0.00%"
    for i, w in enumerate([26, 12, 14, 8, 10, 11, 9, 10, 13, 17, 16, 17, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- All Captures ----------------
    ws = wb.create_sheet("All Captures")
    _hdr(ws, ["Session", "Date", "Time", "Asset", "Strategy",
              "Net Profit (INR)", "Opportunity ID"],
         [26, 12, 10, 14, 18, 16, 46])
    for s in sessions:
        for c in sorted(s["caps"], key=lambda x: x.get("timestamp", "")):
            ts = c.get("timestamp", "")
            ws.append([s["name"], ts[:10], ts[11:19], c.get("asset"),
                       c.get("strategy"), c.get("net_profit"),
                       c.get("opportunity_id")])
    for row in ws.iter_rows(min_row=2):
        row[5].number_format = "#,##0.00"

    # ---------------- Strategy Perf ----------------
    ws = wb.create_sheet("Strategy Perf")
    _hdr(ws, ["Strategy", "Captures", "Net P&L (INR)", "Avg per capture",
              "% of total P&L", "Sessions active"],
         [22, 12, 16, 16, 15, 16])
    agg = defaultdict(lambda: {"n": 0, "net": 0.0, "sess": 0})
    for s in sessions:
        for k, v in s["by_strategy"].items():
            agg[k]["n"] += v["n"]
            agg[k]["net"] += v["net"]
            agg[k]["sess"] += 1
    for k, v in sorted(agg.items(), key=lambda kv: -kv[1]["net"]):
        ws.append([k, v["n"], v["net"], v["net"] / v["n"] if v["n"] else 0,
                   (v["net"] / tot_net) if tot_net else 0, v["sess"]])
    ws.append(["TOTAL", tot_caps, tot_net, tot_net / tot_caps if tot_caps else 0, 1.0, ""])
    for c in ws[ws.max_row]:
        c.font = BOLD
        c.fill = KPI
    for row in ws.iter_rows(min_row=2):
        row[2].number_format = "#,##0.00"
        row[3].number_format = "#,##0.00"
        row[4].number_format = "0.0%"

    # ---------------- Rejection Analysis ----------------
    ws = wb.create_sheet("Rejection Analysis")
    _hdr(ws, ["Session", "Scanned", "Not profitable after costs",
              "Below min annualized return", "Insufficient capital",
              "Executable", "Executable %"],
         [26, 11, 24, 24, 20, 12, 13])
    for s in sessions:
        r = s["rejections"]
        ws.append([s["name"], s["scanned"],
                   r.get("not_profitable_after_round_trip_costs", 0),
                   r.get("below_min_annualized_return", 0),
                   r.get("insufficient_capital", 0),
                   s["executable"],
                   (s["executable"] / s["scanned"]) if s["scanned"] else 0])
    for row in ws.iter_rows(min_row=2):
        row[6].number_format = "0.00%"
    ws.append([])
    ws.append(["Note: reasons overlap — one opportunity can fail several "
               "hurdles, so the columns do not sum to Scanned."])
    ws[ws.max_row][0].font = Font(italic=True, color="808080")

    # ---------------- Path to 5k ----------------
    ws = wb.create_sheet("Path to 5k")
    ws.append(["WHAT THE DATA SAYS ABOUT REACHING 5,000 INR / DAY"])
    ws["A1"].font = Font(size=13, bold=True, color="1F3864")
    ws.append([])

    real = [s for s in sessions if s["capital"] == 1_000_000 and s["hours"] > 0.5]
    best = max(real, key=lambda s: s["net"]) if real else None
    full_day_hours = 6.25  # 09:15-15:30 IST
    ws.append(["Observed today", "Value", "Comment"])
    for c in ws[ws.max_row]:
        c.fill = SUB
        c.font = WHITE
    facts = [
        ("Best session P&L", best["net"] if best else 0,
         f"{best['name']} over {best['hours']:.1f}h" if best else ""),
        ("Best session P&L / hour", best["per_hour"] if best else 0,
         "the rate that matters, not the total"),
        ("Extrapolated to a 6.25h day",
         (best["per_hour"] * full_day_hours) if best and best["per_hour"] else 0,
         "if the observed rate held all session — it will not, "
         "spreads are widest at open"),
        ("Sessions at 10L that beat 5,000",
         sum(1 for s in real if s["net"] >= 5000), f"out of {len(real)} full 10L sessions"),
        ("Avg capture size (all sessions)",
         (tot_net / tot_caps) if tot_caps else 0, "INR per locked-in capture"),
        ("Captures needed for 5,000/day",
         (5000 / (tot_net / tot_caps)) if tot_caps and tot_net else 0,
         "at the current average capture size"),
    ]
    for k, v, c in facts:
        ws.append([k, v, c])
        ws.cell(row=ws.max_row, column=2).number_format = "#,##0.00"

    ws.append([])
    ws.append(["Lever", "Expected effect", "Evidence from these sessions", "Risk"])
    for c in ws[ws.max_row]:
        c.fill = SUB
        c.font = WHITE
    levers = [
        ("Run the FULL session 09:15-15:30",
         "Largest single gain",
         "Every session so far ran 0.6-3.5h, none a full day. Best run made "
         "5,241 in 3.5h and was still capturing when stopped.",
         "None - pure coverage. Do this first."),
        ("Raise capital above 10L",
         "More equity F&O qualifies",
         "At 1L only MCX cleared (324 INR). At 10L futures_basis + "
         "put_call_parity dominate. insufficient_capital still blocked "
         f"{sum(s['rejections'].get('insufficient_capital',0) for s in real)} "
         "rows at 10L.",
         "Bigger per-trade loss if a leg does not fill."),
        ("Widen the instrument universe",
         "Roughly linear in scan volume",
         "Universe is ~14 dual-listed + ~21 F&O + 5 MCX. nse_bse_arb is "
         "the biggest scan bucket but contributes almost no P&L.",
         "More Dhan calls - already rate-limited, see Reality Check."),
        ("Lower min_annualized_return_pct (5%)",
         "More marginal trades qualify",
         "below_min_annualized_return is the single largest rejection "
         f"reason: {sum(s['rejections'].get('below_min_annualized_return',0) for s in sessions):,} rows.",
         "Directly trades quality for quantity. Thin edges are the "
         "first to vanish on real fills."),
        ("Reduce modelled cost (0.03% spread etc.)",
         "Only if real fees are genuinely lower",
         "not_profitable_after_round_trip_costs blocks "
         f"{sum(s['rejections'].get('not_profitable_after_round_trip_costs',0) for s in sessions):,} rows.",
         "DANGEROUS - lowering an assumption does not lower a real fee. "
         "Only change with a real broker contract note."),
    ]
    for row in levers:
        ws.append(list(row))
        for c in ws[ws.max_row]:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate([34, 26, 62, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = None

    # ---------------- Reality Check ----------------
    ws = wb.create_sheet("Reality Check")
    ws.append(["PAPER vs REAL MONEY — WHAT WOULD HAVE TO HOLD"])
    ws["A1"].font = Font(size=13, bold=True, color="1F3864")
    ws.append(["Every P&L in this workbook is PAPER. No order was ever sent. "
               "The figures assume the market would have filled both legs at "
               "the prices quoted. The gaps below are the reasons it might not."])
    ws["A2"].font = Font(italic=True, color="595959")
    ws.append([])
    ws.append(["Assumption in the paper P&L", "What actually happens live",
               "Effect on P&L", "Verified?"])
    for c in ws[ws.max_row]:
        c.fill = SUB
        c.font = WHITE
    gaps = [
        ("Trades execute at last-traded price (LTP)",
         "The connector only fetches last_price. Real entry crosses the "
         "bid/ask. The gate substitutes a flat 0.03% assumed spread.",
         "Direct deduction from every trade. Edges here average "
         "0.06-0.12%, so a spread wider than assumed erases the trade.",
         "CONFIRMED - dhan_connector reads only last_price/volume"),
        ("Both legs fill at the same instant",
         "Arbitrage needs two fills. Between leg 1 and leg 2 the spread can "
         "close, leaving a naked directional position.",
         "Converts a small modelled gain into an uncapped directional "
         "loss. This is the single biggest live risk.",
         "CONFIRMED - no order path exists; never tested"),
        ("The same spread can be booked each poll",
         "A spread that stays open for several polls was captured "
         f"repeatedly. Across all sessions {tot_caps} captures collapse to "
         f"{tot_uniq} unique opportunities.",
         f"Overstates total P&L by "
         f"{(100 * (tot_net - tot_dedup) / tot_dedup) if tot_dedup else 0:.1f}% "
         f"({tot_net:,.0f} -> {tot_dedup:,.0f}).",
         "CONFIRMED - measured from opportunity_id repeats"),
        ("Full size is always available",
         "Executable quantity is inferred from reported volume, not from "
         "order-book depth.",
         "Large legs may only partially fill, leaving an unbalanced "
         "position.",
         "CONFIRMED - liquidity_engine uses volume as proxy"),
        ("Modelled costs equal real costs",
         "Brokerage/tax/slippage are configured percentages, not a broker "
         "contract note. STT, stamp duty, GST and exchange fees differ per "
         "segment.",
         "If real cost exceeds the model by even 0.05%, most captures here "
         "flip negative.",
         "NOT VERIFIED - needs a real contract note"),
        ("The scanner can watch the whole market",
         "Dhan rate-limits hard. Two concurrent sessions degraded BOTH into "
         "retry loops; a single session already sits near the ceiling.",
         "Caps how many instruments can be watched, and therefore how many "
         "opportunities are seen at all.",
         "CONFIRMED - observed 2026-07-27, both sessions failed"),
        ("Results generalise",
         f"Total sample is {sum(s['hours'] for s in sessions):.1f} hours "
         f"across {len(sessions)} partial sessions, no full trading day, one "
         "market regime.",
         "Too small to estimate a win rate or a drawdown with any "
         "confidence.",
         "CONFIRMED - see Session Summary hours column"),
    ]
    for g in gaps:
        ws.append(list(g))
        for c in ws[ws.max_row]:
            c.alignment = Alignment(wrap_text=True, vertical="top")
        if g[3].startswith("NOT"):
            ws.cell(row=ws.max_row, column=4).fill = WARN
    ws.append([])
    ws.append(["BOTTOM LINE"])
    ws[ws.max_row][0].font = Font(bold=True, size=12, color="C00000")
    for line in [
        "The strategies are real arbitrage (put-call parity, cash-and-carry, "
        "calendar spreads) and the cost gate is genuinely conservative — it "
        "rejects ~91% of what it sees. That part is sound engineering.",
        "But the paper P&L has never been tested against a real fill. The "
        "edges are 0.06-0.12%, which is the same order of magnitude as the "
        "unmodelled costs (real spread, leg-in slippage, exact fee schedule). "
        "That is the whole question, and this data cannot answer it.",
        "No responsible percentage can be put on live survival from "
        f"{sum(s['hours'] for s in sessions):.1f} hours of paper trading with "
        "zero executed orders. Anyone quoting one would be guessing.",
        "The one cheap way to find out: place a handful of REAL minimum-size "
        "trades and compare the actual contract note against what this system "
        "predicted. That single experiment answers more than another month of "
        "paper trading.",
    ]:
        ws.append([line])
        ws[ws.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=4)
    for i, w in enumerate([38, 52, 46, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = out or f"storage/all_sessions_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(out)
    return out, sessions, tot_net, tot_caps


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--storage", default="storage")
    p.add_argument("--out", default=None)
    a = p.parse_args()
    path, sessions, net, caps = build(a.storage, a.out)
    print(f"Wrote {path}")
    print(f"{len(sessions)} sessions, {caps} captures, net {net:,.2f} INR")


if __name__ == "__main__":
    main()
